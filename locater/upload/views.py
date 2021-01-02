from django.shortcuts import render,redirect
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse
import os,requests
import openpyxl
from dotenv import load_dotenv
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(BASE_DIR,'.env'))
api_key = os.environ['API_KEY']


# Create your views here.
def upload_xls(request):
    if request.method == "GET":
        return render(request,'upload.html',{'message':''})
    
    elif request.method == "POST":
        try:
            uploaded_file = request.FILES['file']
            fs = FileSystemStorage()
            uploaded_file = fs.save(uploaded_file.name,uploaded_file)
            uploaded_file_on_server = os.path.join(BASE_DIR,'media',uploaded_file)
            work_book = openpyxl.load_workbook(uploaded_file_on_server)
            for worksheet in work_book.worksheets:
                row = 2
                col = 1
                while row <= worksheet.max_row:
                    if worksheet.cell(row,col).value != "":
                        address = worksheet.cell(row,col).value
                        # print("address----->",address)
                        try:
                            try:
                                response = requests.get("http://www.mapquestapi.com/geocoding/v1/address?key={}&location={}".format(api_key,address))
                                data = response.json()
                            except Exception:
                                return render(request,"upload.html",{"message":"Error in fetching Lat/Lng from mapquest"})
                            latitude=data["results"][0]["locations"][0]["latLng"]["lat"]
                            longitude=data["results"][0]["locations"][0]["latLng"]["lng"]
                            worksheet.cell(row,col+1).value = latitude
                            worksheet.cell(row,col+2).value = longitude
                            row+=1
                        except Exception as err:
                            return render(request,'upload.html',{"message":err})

            
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="{}"'.format(uploaded_file)
            work_book.save(response)
            return response
        except Exception as err:
            return render(request,"upload.html",{"message":err})